# services/eta_export.py

from typing import List, Dict, Iterable, Tuple, Callable, Optional
import io, csv, re
from datetime import datetime
from zoneinfo import ZoneInfo  # Py3.9+
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

PREFERRED_COLS = [
    "RefNo", "DateScheduled", "ProductionStatus", "ProductionLine",
    "InventoryItem", "Descn", "Instance", "FixedLine"
]

# services/eta_export.py (top-level, near DISPLAY_HEADERS)
_COST_COL_RE = re.compile(r"(^|_|-|\b)cost(s|ed|ing)?(\b|_|-|$)", re.IGNORECASE)


# --- NEW: headers for exactly what the table shows ---
DISPLAY_HEADERS = ["Group", "Item", "Status"]


def _strip_cost_columns(headers: List[str]) -> List[str]:
    return [h for h in headers if not _COST_COL_RE.search(h)]


def ordered_headers(rows: Iterable[Dict]) -> List[str]:
    rows = list(rows)
    if not rows:
        return PREFERRED_COLS[:]
    seen: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.append(k)
    # Build in preferred order, then strip any cost-like columns
    headers = [c for c in PREFERRED_COLS if c in seen] + [c for c in seen if c not in PREFERRED_COLS]
    headers = [h for h in headers if not _COST_COL_RE.search(h)]
    return headers


def apply_filters(rows: Iterable[Dict], *, status: str = "", group: str = "", supplier: str = "") -> List[Dict]:
    s = (status or "").strip().lower()
    g = (group or "").strip().lower()
    sup = (supplier or "").strip().upper()

    def ok(r: Dict) -> bool:
        if s and (str(r.get("ProductionStatus", "")).strip().lower() != s):
            return False
        if g and (str(r.get("ProductionLine", "")).strip().lower() != g):
            return False
        if sup and (str(r.get("Instance", "")).strip().upper() != sup):
            return False
        return True

    return [r for r in rows if ok(r)]


# --- NEW: turn full rows into the exact 3 columns you render in the table ---
def map_rows_to_display(rows: Iterable[Dict]) -> Tuple[List[Dict], List[str]]:
    out: List[Dict] = []
    for r in rows:
        group = (r.get("ProductionLine") or "").strip()
        inv   = (r.get("InventoryItem") or "").strip()
        fixed = (r.get("FixedLine") or "").strip()
        item  = f"{inv} ({fixed})" if fixed else inv
        status = (r.get("ProductionStatus") or "").strip()
        out.append({"Group": group, "Item": item, "Status": status})
    return out, DISPLAY_HEADERS


def safe_base_filename(name_or_id: str, tz: str = "Australia/Sydney") -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", (name_or_id or "")).strip("-") or "report"
    today = datetime.now(ZoneInfo(tz)).date().isoformat()
    return f"{base}-open-orders-{today}"


def to_csv_bytes(rows: Iterable[Dict], headers: List[str]) -> bytes:
    headers = _strip_cost_columns(headers)
    sio = io.StringIO()
    w = csv.writer(sio, lineterminator="\n")
    w.writerow(headers)
    for r in rows:
        w.writerow([r.get(h, "") for h in headers])
    return sio.getvalue().encode("utf-8-sig")  # BOM for Excel


def to_excel_bytes(rows: Iterable[Dict], headers: List[str]) -> bytes:
    headers = _strip_cost_columns(headers)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    max_widths = [len(str(h)) for h in headers]
    for r in rows:
        row_vals = [r.get(h, "") for h in headers]
        ws.append(row_vals)
        for i, v in enumerate(row_vals):
            L = len(str(v)) if v is not None else 0
            if L > max_widths[i]:
                max_widths[i] = L
    for idx, w in enumerate(max_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(w + 2, 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def fetch_report_rows_and_name(
    obfuscated_id: str,
    *,
    query_db: Callable,
    get_db: Callable,
    get_open_orders: Callable,
    get_open_orders_by_group: Callable
) -> Tuple[Optional[List[Dict]], Optional[str]]:
    customer = query_db(
        "SELECT dd_name, cbr_name, field_type FROM customers WHERE obfuscated_id = ?",
        (obfuscated_id,), one=True
    )
    if not customer:
        return None, None

    dd_name, cbr_name, field_type = customer
    if field_type == "Customer Group":
        data_dd  = get_open_orders_by_group(get_db(), dd_name, "DD") if dd_name else []
        data_cbr = get_open_orders_by_group(get_db(), cbr_name, "CBR") if cbr_name else []
    else:
        data_dd  = get_open_orders(get_db(), dd_name, "DD") if dd_name else []
        data_cbr = get_open_orders(get_db(), cbr_name, "CBR") if cbr_name else []

    combined = (data_dd or []) + (data_cbr or [])

    if (dd_name == cbr_name) or (not cbr_name):
        customer_name = dd_name or cbr_name or ""
    elif not dd_name:
        customer_name = cbr_name
    else:
        customer_name = f"{dd_name} & {cbr_name}"

    return combined, customer_name
