# services/export.py
from typing import List, Dict, Iterable, Tuple, Callable, Optional
import io, csv, re
from datetime import datetime
from zoneinfo import ZoneInfo  # Py3.9+
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from typing import Any


DROP_KEY_RE = re.compile(
    r"(cost|buy|cogs|margin|markup|wholesale|supplier.?price|pkid\b)",
    re.I,
)
PREFERRED_COLS = [
    "RefNo", "DateScheduled", "ProductionStatus", "ProductionLine",
    "InventoryItem", "Descn", "Instance", "FixedLine"
]
DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def _sanitize_for_excel(value: object) -> str:
    s = "" if value is None else str(value)
    return "'" + s if s.startswith(DANGEROUS_PREFIXES) else s


def scrub_sensitive(rows):
    safe = []
    for r in rows:
        safe.append({
            k: v
            for k, v in r.items()
            if not DROP_KEY_RE.search(str(k))
        })
    return safe


def ordered_headers(rows: Iterable[Dict[str, Any]]) -> List[str]:
    rows = list(rows)
    if not rows:
        return PREFERRED_COLS[:]
    seen: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.append(k)
    return [c for c in PREFERRED_COLS if c in seen] + [c for c in seen if c not in PREFERRED_COLS]


def apply_filters(rows: Iterable[Dict[str, Any]], *, status: str = "", group: str = "", supplier: str = "") -> List[Dict[str, Any]]:
    s = (status or "").strip().lower()
    g = (group or "").strip().lower()
    sup = (supplier or "").strip().upper()

    def ok(r: Dict) -> bool:
        if s and (str(r.get("ProductionStatus","")).strip().lower() != s): return False
        if g and (str(r.get("ProductionLine","")).strip().lower() != g): return False
        if sup and (str(r.get("Instance","")).strip().upper() != sup): return False
        return True

    return [r for r in rows if ok(r)]


def safe_base_filename(name_or_id: str, tz: str = "Australia/Sydney") -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", (name_or_id or "")).strip("-") or "report"
    today = datetime.now(ZoneInfo(tz)).date().isoformat()
    return f"{base}-open-orders-{today}"


def to_csv_bytes(rows: Iterable[Dict[str, Any]], headers: List[str]) -> bytes:
    sio = io.StringIO()
    w = csv.writer(sio, lineterminator="\n")
    w.writerow(headers)
    for r in rows:
        w.writerow([_sanitize_for_excel(r.get(h, "")) for h in headers])
    return sio.getvalue().encode("utf-8-sig")


def to_excel_bytes(rows: Iterable[Dict[str, Any]], headers: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    if not headers:
        return to_csv_bytes(rows, [])

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    max_widths = [len(str(h)) for h in headers]
    for r in rows:
        row_vals = [_sanitize_for_excel(r.get(h, "")) for h in headers]
        ws.append(row_vals)
        for i, v in enumerate(row_vals):
            L = len(v)
            if L > max_widths[i]:
                max_widths[i] = L
    for idx, w_ in enumerate(max_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(w_ + 2, 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def fetch_report_rows_and_name(
    obfuscated_id: str,
    *,
    query_db: Callable,
    get_db: Callable,
    get_open_orders: Callable,
    get_open_orders_by_group: Callable
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
    """Pure function wrapper around your existing data access."""
    customer = query_db(
        "SELECT dd_name, cbr_name, field_type FROM customers WHERE obfuscated_id = ?",
        (obfuscated_id,),
        one=True
    )
    if not customer:
        return None, None

    dd_name, cbr_name, field_type = customer
    if field_type == "Customer Group":
        data_dd  = get_open_orders_by_group(get_db(), dd_name, "DD") if dd_name else []
        data_cbr = get_open_orders_by_group(get_db(), cbr_name, "CBR") if cbr_name else []
    else:
        data_dd  = get_open_orders(get_db(), dd_name, "DD") if dd_name else []
        data_cbr = get_open_orders(get_db(), cbr_name, "CBR") if cbr_name else []

    combined = (data_dd or []) + (data_cbr or [])

    if (dd_name == cbr_name) or (not cbr_name):
        customer_name = dd_name or cbr_name or ""
    elif not dd_name:
        customer_name = cbr_name
    else:
        customer_name = f"{dd_name} & {cbr_name}"

    return combined, customer_name
