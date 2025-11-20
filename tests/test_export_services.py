# tests/test_export.py
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Tuple
import pytest

from openpyxl import load_workbook

import services.export as export


# ---------------------------
# ordered_headers
# ---------------------------

def test_ordered_headers_prefers_known_cols_first_then_new_in_seen_order():
    rows = [
        {
            "InventoryItem": "X",
            "Descn": "Y",
            "Foo": "1",
            "Bar": "2",
            "RefNo": "ABC",
        },
        {
            "Zed": "3",
            "ProductionStatus": "Open",
            "Baz": "4",
        },
    ]
    headers = export.ordered_headers(rows)

    # Preferred columns should appear first, but only those that are present
    preferred_present = [c for c in export.PREFERRED_COLS if c in {"InventoryItem", "Descn", "RefNo", "ProductionStatus"}]
    assert headers[: len(preferred_present)] == preferred_present

    # All seen keys should appear exactly once
    for k in ["InventoryItem", "Descn", "Foo", "Bar", "RefNo", "Zed", "ProductionStatus", "Baz"]:
        assert k in headers
    assert len(headers) == len(set(headers))


def test_ordered_headers_when_rows_empty_returns_copy_of_preferred_cols():
    headers = export.ordered_headers([])
    assert headers == export.PREFERRED_COLS
    assert headers is not export.PREFERRED_COLS  # copy, not same object


# ---------------------------
# apply_filters
# ---------------------------

def _row(ps="", pl="", inst=""):
    return {
        "ProductionStatus": ps,
        "ProductionLine": pl,
        "Instance": inst,
    }


def test_apply_filters_matches_case_insensitively_and_trims():
    rows = [
        _row("Open", "Cutting", "dd"),
        _row("open", "CUTTING", "DD"),
        _row("Closed", "Cutting", "DD"),
    ]
    out = export.apply_filters(rows, status=" open ", group=" cutting ", supplier=" dd ")
    # Only rows 0 and 1 match all three criteria; 2 fails status
    assert out == rows[:2]


def test_apply_filters_with_some_empty_filters():
    rows = [
        _row("Open", "Cutting", "CBR"),
        _row("Open", "Sewing", "DD"),
    ]
    # Filter just by group
    out = export.apply_filters(rows, group=" sewing ")
    assert out == [rows[1]]


def test_apply_filters_no_filters_returns_all():
    rows = [_row("Open", "A", "X"), _row("Closed", "B", "Y")]
    assert export.apply_filters(rows) == rows


# ---------------------------
# safe_base_filename
# ---------------------------

def test_safe_base_filename_sanitises_and_adds_date_suffix():
    fname = export.safe_base_filename("ACME Pty Ltd / East|West")
    # e.g. "ACME-Pty-Ltd-East-West-open-orders-YYYY-MM-DD"
    assert fname.startswith("ACME-Pty-Ltd-East-West-open-orders-")
    assert re.match(r".+-\d{4}-\d{2}-\d{2}$", fname)


def test_safe_base_filename_empty_input_uses_report():
    fname = export.safe_base_filename("")
    assert fname.startswith("report-open-orders-")
    assert re.match(r"report-open-orders-\d{4}-\d{2}-\d{2}$", fname)


# ---------------------------
# to_csv_bytes
# ---------------------------

def test_to_csv_bytes_writes_bom_and_rows_in_header_order():
    headers = ["A", "B", "C"]
    rows = [{"A": 1, "B": 2, "C": 3}, {"A": "x", "B": "y", "C": "z"}]
    b = export.to_csv_bytes(rows, headers)

    # Excel-friendly UTF-8 BOM
    assert b.startswith(b"\xef\xbb\xbf")

    s = b.decode("utf-8-sig")
    lines = s.splitlines()
    assert lines[0] == "A,B,C"
    assert lines[1] == "1,2,3"
    assert lines[2] == "x,y,z"


# ---------------------------
# to_excel_bytes
# ---------------------------

def test_to_excel_bytes_builds_basic_workbook_with_header_styles_and_freeze():
    headers = ["RefNo", "InventoryItem", "Descn"]
    rows = [
        {"RefNo": "R1", "InventoryItem": "ItemA", "Descn": "Alpha"},
        {"RefNo": "R2", "InventoryItem": "ItemB", "Descn": "Beta"},
    ]
    data = export.to_excel_bytes(rows, headers)

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    assert ws.title == "Report"
    # Headers
    assert [c.value for c in ws[1]] == headers
    # Bold font on headers
    for c in ws[1]:
        assert c.font.bold

    # Freeze panes at A2
    assert ws.freeze_panes == "A2"

    # Data rows
    assert [cell.value for cell in ws[2]] == ["R1", "ItemA", "Alpha"]
    assert [cell.value for cell in ws[3]] == ["R2", "ItemB", "Beta"]

    # Column widths should be > header length and <= 60
    for dim in ws.column_dimensions.values():
        if dim.width is not None:
            assert dim.width <= 60


def test_to_excel_bytes_with_no_headers_falls_back_to_csv_bytes():
    rows = [{"A": 1}]
    excel_fallback = export.to_excel_bytes(rows, headers=[])
    csv_equiv = export.to_csv_bytes(rows, headers=[])
    assert excel_fallback == csv_equiv


# ---------------------------
# fetch_report_rows_and_name
# ---------------------------

class _StubDB:
    pass


@pytest.fixture
def stubs():
    """Common stubs for DB-related callables."""
    db = _StubDB()

    def get_db():
        return db

    def make_orders(prefix: str) -> List[Dict[str, Any]]:
        return [
            {"RefNo": f"{prefix}-1", "Instance": prefix},
            {"RefNo": f"{prefix}-2", "Instance": prefix},
        ]

    def get_open_orders(conn, name: str, inst: str) -> List[Dict[str, Any]]:
        assert conn is db
        return make_orders(inst)

    def get_open_orders_by_group(conn, name: str, inst: str) -> List[Dict[str, Any]]:
        assert conn is db
        return [{"RefNo": f"{inst}-G-1", "Instance": inst}]

    return get_db, get_open_orders, get_open_orders_by_group


def test_fetch_report_rows_and_name_returns_none_when_customer_missing(stubs):
    get_db, get_open_orders, get_open_orders_by_group = stubs

    def query_db(sql: str, params: Tuple[Any, ...], one: bool = False):
        return None  # no customer

    rows, name = export.fetch_report_rows_and_name(
        "obf123",
        query_db=query_db,
        get_db=get_db,
        get_open_orders=get_open_orders,
        get_open_orders_by_group=get_open_orders_by_group,
    )
    assert rows is None and name is None


def test_fetch_report_rows_and_name_individual_customer_uses_get_open_orders(stubs):
    get_db, get_open_orders, get_open_orders_by_group = stubs

    def query_db(sql: str, params: Tuple[Any, ...], one: bool = False):
        # dd_name, cbr_name, field_type, display_name
        return ("ACME DD", "ACME CBR", "Customer", None)

    rows, name = export.fetch_report_rows_and_name(
        "obf123",
        query_db=query_db,
        get_db=get_db,
        get_open_orders=get_open_orders,
        get_open_orders_by_group=get_open_orders_by_group,
    )

    # Combined DD + CBR
    assert isinstance(rows, list) and len(rows) == 4
    # Name combines when both present and different
    assert name == "ACME DD & ACME CBR"


def test_fetch_report_rows_and_name_group_customer_uses_group_accessor(stubs):
    get_db, get_open_orders, get_open_orders_by_group = stubs

    def query_db(sql: str, params: Tuple[Any, ...], one: bool = False):
        return ("Grp DD", "Grp CBR", "Customer Group", None)

    rows, name = export.fetch_report_rows_and_name(
        "obf456",
        query_db=query_db,
        get_db=get_db,
        get_open_orders=get_open_orders,  # should not be used here
        get_open_orders_by_group=get_open_orders_by_group,
    )

    # Group accessor returns 1 per instance in our stub => 2 total
    assert len(rows) == 2
    assert {r["Instance"] for r in rows} == {"DD", "CBR"}
    assert name == "Grp DD & Grp CBR"


def test_fetch_report_rows_and_name_handles_same_dd_cbr_name(stubs):
    get_db, get_open_orders, get_open_orders_by_group = stubs

    def query_db(sql: str, params: Tuple[Any, ...], one: bool = False):
        return ("SameName", "SameName", "Customer", None)

    rows, name = export.fetch_report_rows_and_name(
        "obf789",
        query_db=query_db,
        get_db=get_db,
        get_open_orders=get_open_orders,
        get_open_orders_by_group=get_open_orders_by_group,
    )

    assert len(rows) == 4
    assert name == "SameName"  # avoids duplicate "A & A"


def test_fetch_report_rows_and_name_handles_missing_one_side(stubs):
    get_db, get_open_orders, get_open_orders_by_group = stubs

    def query_db(sql: str, params: Tuple[Any, ...], one: bool = False):
        # No CBR side
        return ("OnlyDD", "", "Customer", None)

    rows, name = export.fetch_report_rows_and_name(
        "obf999",
        query_db=query_db,
        get_db=get_db,
        get_open_orders=get_open_orders,
        get_open_orders_by_group=get_open_orders_by_group,
    )

    # Only DD orders (2 in our stub)
    assert len(rows) == 2
    assert name == "OnlyDD"
