from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Tuple
from openpyxl import load_workbook
import pytest

from services import export as export_mod


@pytest.fixture
def app_module():
    """Import app module after environment is set up."""
    import app as app_module
    return app_module


def _rows_sample() -> List[Dict[str, Any]]:
    return [
        {
            "RefNo": "R1",
            "DateScheduled": "2025-10-20",
            "ProductionStatus": "Open",
            "ProductionLine": "Cutting",
            "InventoryItem": "ITEM1",
            "Descn": "Alpha",
            "Instance": "DD",
            "FixedLine": "GLIDE",
        },
        {
            "RefNo": "R2",
            "DateScheduled": "2025-10-21",
            "ProductionStatus": "Open",
            "ProductionLine": "Sewing",
            "InventoryItem": "ITEM2",
            "Descn": "Beta",
            "Instance": "CBR",
            "FixedLine": "ROLLER",
        },
    ]


def test_download_csv_happy_path(client, monkeypatch, app_module):
    # Stub data fetcher: returns rows + customer name
    monkeypatch.setattr(
        app_module,
        "fetch_report_rows_and_name",
        lambda *a, **k: (_rows_sample(), "Acme Widgets"),
    )

    # Let filters/order/export use real implementations
    resp = client.get("/abcd123/download.csv?statusFilter=open&group=cutting&supplier=dd")
    assert resp.status_code == 200
    assert resp.mimetype.startswith("text/csv")

    # Filename pattern includes date; don't depend on today exactly
    cd = resp.headers.get("Content-Disposition", "")
    # attachment; filename=Acme-Widgets-open-orders-YYYY-MM-DD.csv OR quoted
    assert "attachment;" in cd
    assert cd.endswith(".csv") or cd.endswith('.csv"')
    assert re.search(r"Acme-Widgets-open-orders-\d{4}-\d{2}-\d{2}\.csv", cd) is not None

    # CSV content basic checks
    data = resp.data
    assert data.startswith(b"\xef\xbb\xbf")  # BOM for Excel
    text = data.decode("utf-8-sig")
    lines = text.splitlines()
    expected_headers = export_mod.ordered_headers(_rows_sample())
    assert lines[0] == ",".join(expected_headers)
    # First row values align with headers
    first_vals = [_rows_sample()[0].get(h, "") for h in expected_headers]
    assert lines[1] == ",".join(str(v) for v in first_vals)


def test_download_xlsx_happy_path(client, monkeypatch, app_module):
    monkeypatch.setattr(
        app_module,
        "fetch_report_rows_and_name",
        lambda *a, **k: (_rows_sample(), "Acme Widgets"),
    )

    resp = client.get("/abcd123/download.xlsx?status=open&group=cutting&supplier=DD")
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Filename sanity
    cd = resp.headers.get("Content-Disposition", "")
    assert "attachment;" in cd
    assert cd.endswith(".xlsx") or cd.endswith('.xlsx"')
    assert re.search(r"Acme-Widgets-open-orders-\d{4}-\d{2}-\d{2}\.xlsx", cd) is not None

    # Load workbook to ensure it's valid and headers present
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    assert ws.title == "Report"
    headers = [c.value for c in ws[1]]
    assert headers == export_mod.ordered_headers(_rows_sample())
    # First data row matches
    assert [c.value for c in ws[2]] == [_rows_sample()[0].get(h, "") for h in headers]


def test_download_404_when_customer_missing(client, monkeypatch, app_module):
    monkeypatch.setattr(app_module, "fetch_report_rows_and_name", lambda *a, **k: (None, None))
    resp = client.get("/nope/download.csv")
    assert resp.status_code == 404


def test_download_400_when_bad_format(client, monkeypatch, app_module):
    monkeypatch.setattr(
        app_module,
        "fetch_report_rows_and_name",
        lambda *a, **k: (_rows_sample(), "Acme Widgets"),
    )
    resp = client.get("/abcd123/download.pdf")
    assert resp.status_code == 400


def test_filter_precedence_status_and_legacy_params(client, monkeypatch, app_module):
    # Capture what apply_filters receives
    seen_kwargs = {}

    def spy_apply_filters(rows, **kwargs):
        seen_kwargs.update(kwargs)
        return rows  # pass-through

    monkeypatch.setattr(
        app_module,
        "fetch_report_rows_and_name",
        lambda *a, **k: (_rows_sample(), "Acme Widgets"),
    )
    monkeypatch.setattr(app_module, "apply_filters", spy_apply_filters)

    resp = client.get(
        "/abcd123/download.csv?status=primary&statusFilter=secondary&groupFilter=G&supplierFilter=S"
    )
    assert resp.status_code == 200

    # Precedence: "status" should win over "statusFilter"
    assert seen_kwargs["status"] == "primary"
    # Only legacy names exist for group/supplier in this request
    assert seen_kwargs["group"] == "G"
    assert seen_kwargs["supplier"] == "S"


def test_report_render_preserves_obfuscated_id(client, monkeypatch):
    job_id = "job123"
    result = {
        "template": "report.html",
        "status": 200,
        "context": {"obfuscated_id": "obf-abc123", "some": "ctx"}
    }
    monkeypatch.setattr("app.get_job", lambda _: {"status": "completed", "result": result}, raising=True)

    r = client.get(f"/report/{job_id}")
    assert r.status_code == 200
    assert b'data-obf="obf-abc123"' in r.data  # comes from template div
