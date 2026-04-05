"""Tests for edge cases: empty downloads, malformed API input, zero-row exports."""
import io
import json
import pytest
from openpyxl import load_workbook

from services.export import to_csv_bytes, to_excel_bytes, ordered_headers, scrub_sensitive


# ---------- Empty/zero-row exports ----------

class TestEmptyExports:
    def test_csv_zero_rows(self):
        """CSV with headers but no data rows."""
        headers = ["RefNo", "Status"]
        data = to_csv_bytes([], headers)
        text = data.decode("utf-8-sig")
        lines = text.strip().splitlines()
        assert len(lines) == 1  # header only
        assert lines[0] == "RefNo,Status"

    def test_csv_no_headers_no_rows(self):
        data = to_csv_bytes([], [])
        text = data.decode("utf-8-sig")
        # Just a BOM + empty header line
        assert text.strip() == ""

    def test_excel_zero_rows(self):
        """XLSX with headers but no data rows."""
        headers = ["RefNo", "Status"]
        data = to_excel_bytes([], headers)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Header row exists
        assert [c.value for c in ws[1]] == headers
        # No data row
        assert ws[2][0].value is None

    def test_ordered_headers_with_empty_rows(self):
        """ordered_headers returns preferred cols for empty input."""
        assert ordered_headers([]) == [
            "RefNo", "DateScheduled", "ProductionStatus", "ProductionLine",
            "InventoryItem", "Descn", "Instance", "FixedLine",
        ]

    def test_scrub_sensitive_empty(self):
        assert scrub_sensitive([]) == []


# ---------- Malformed API input ----------

class TestMalformedApiInput:
    def test_create_customer_no_json_body(self, client, api_headers):
        """POST with no body at all."""
        r = client.post(
            "/api/v1/customers",
            headers=api_headers,
            data="",
            content_type="application/json",
        )
        assert r.status_code == 400

    def test_create_customer_invalid_json(self, client, api_headers):
        """POST with invalid JSON."""
        r = client.post(
            "/api/v1/customers",
            headers=api_headers,
            data="not json",
            content_type="application/json",
        )
        # get_json(silent=True) returns None, so body becomes {}
        # Then validation fails: no dd_name or cbr_name
        assert r.status_code == 400

    def test_create_customer_empty_object(self, client, api_headers):
        """POST with {} — no names provided → 422 validation error."""
        r = client.post(
            "/api/v1/customers",
            headers=api_headers,
            data=json.dumps({}),
        )
        assert r.status_code == 422
        assert r.get_json()["code"] == "VALIDATION_ERROR"

    def test_create_customer_whitespace_names(self, client, api_headers):
        """POST with only whitespace names — treated as empty → 422."""
        r = client.post(
            "/api/v1/customers",
            headers=api_headers,
            data=json.dumps({"dd_name": "   ", "cbr_name": "  "}),
        )
        assert r.status_code == 422

    def test_update_customer_clear_both_names(self, client, monkeypatch, api_headers):
        """PUT that explicitly clears both names → 422 validation error."""
        from tests.test_api_customers import _make_row
        row = _make_row()
        monkeypatch.setattr(
            "api.customers.query_db",
            lambda *a, **kw: row if kw.get("one") else [row],
            raising=True,
        )
        r = client.put(
            "/api/v1/customers/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",
            headers=api_headers,
            data=json.dumps({"dd_name": "", "cbr_name": ""}),
        )
        assert r.status_code == 422
        assert r.get_json()["code"] == "VALIDATION_ERROR"


# ---------- Download with zero-row results ----------

class TestEmptyDownloads:
    def test_api_download_empty_csv(self, client, monkeypatch, api_headers):
        """Downloading a report with no matching orders returns valid empty CSV."""
        monkeypatch.setattr(
            "api.reports.fetch_report_rows_and_name",
            lambda *a, **kw: ([], "Empty Customer"),
            raising=True,
        )
        r = client.get(
            "/api/v1/reports/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa/download?format=csv",
            headers=api_headers,
        )
        assert r.status_code == 200
        text = r.data.decode("utf-8-sig")
        # Should have at least a header row
        lines = text.strip().splitlines()
        assert len(lines) >= 1

    def test_api_download_empty_xlsx(self, client, monkeypatch, api_headers):
        """Downloading XLSX with no orders returns valid workbook."""
        monkeypatch.setattr(
            "api.reports.fetch_report_rows_and_name",
            lambda *a, **kw: ([], "Empty Customer"),
            raising=True,
        )
        r = client.get(
            "/api/v1/reports/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa/download?format=xlsx",
            headers=api_headers,
        )
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.data))
        assert wb.active is not None

    def test_web_download_empty_csv(self, client, monkeypatch):
        """Web route download with no orders."""
        monkeypatch.setattr(
            "app.fetch_report_rows_and_name",
            lambda *a, **kw: ([], "Empty"),
            raising=True,
        )
        r = client.get("/aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa/download.csv")
        assert r.status_code == 200
